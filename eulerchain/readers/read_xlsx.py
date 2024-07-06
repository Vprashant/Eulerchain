from openpyxl import load_workbook
from typing import Dict, List, Optional
from eulerchain.eulerchain_root.documents.document import Document
from .base_reader import BaseReader


class ReadXlsx(BaseReader):
    def __init__(
            self,
            file_path: str,
            source_column: Optional[str] = None,
            sheet_name: Optional[str] = None,
            encoding: Optional[str] = None,
    ):
        self.file_path = file_path
        self.source_column = source_column
        self.sheet_name = sheet_name
        self.encoding = encoding

    def load(self) -> List[Document]:
        docs = []

        wb = load_workbook(filename=self.file_path, read_only=True, data_only=True)
        ws = wb[self.sheet_name] if self.sheet_name else wb.active

        headers = [cell.value for cell in ws[1]]

        for i, row in enumerate(ws.iter_rows(min_row=2)):
            row_values = [cell.value for cell in row]
            row_dict = dict(zip(headers, row_values))

            content = "\n".join(f"{str(k).strip()}: {str(v).strip()}" for k, v in row_dict.items() if v is not None)
            if self.source_column is not None:
                source = row_dict[self.source_column]
            else:
                source = self.file_path
            metadata = {"source": source, "row": i}
            doc = Document(page_content=content, metadata=metadata)
            docs.append(doc)

        return docs